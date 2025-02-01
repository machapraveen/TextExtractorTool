import pytesseract
from PIL import ImageGrab
import pyautogui
import openpyxl
import tkinter as tk
from tkinter import messagebox
import os
from datetime import datetime
import cv2
import numpy as np

# Configuration
EXCEL_FILE = "Research_Tracker.xlsx"
SHEET_NAME = "Sheet1"
EXTRACTED_TEXT_FOLDER = "extracted_text"
SCREENSHOT_FOLDER = "screenshots"
os.makedirs(EXTRACTED_TEXT_FOLDER, exist_ok=True)
os.makedirs(SCREENSHOT_FOLDER, exist_ok=True)

# Function to draw a rectangle on the screen
def draw_rectangle(start_x, start_y, end_x, end_y):
    """
    Draws a rectangle on the screen to visualize the selected region.
    """
    try:
        # Create a transparent image
        screen_width, screen_height = pyautogui.size()
        img = np.zeros((screen_height, screen_width, 4), dtype=np.uint8)

        # Draw the rectangle
        cv2.rectangle(img, (start_x, start_y), (end_x, end_y), (0, 255, 0, 255), 2)

        # Display the image
        cv2.namedWindow("Region Selection", cv2.WINDOW_NORMAL)
        cv2.setWindowProperty("Region Selection", cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
        cv2.imshow("Region Selection", img)
        cv2.waitKey(1)  # Refresh the window
    except Exception as e:
        print(f"Error drawing rectangle: {e}")

# Function to select a region on the screen
def select_region():
    """
    Allows the user to select a region on the screen by dragging the mouse.
    """
    try:
        print("Select the region by clicking and dragging. Press 'Esc' to cancel.")
        start_x, start_y = pyautogui.position()
        print(f"Starting position: ({start_x}, {start_y})")

        # Draw the initial rectangle
        draw_rectangle(start_x, start_y, start_x, start_y)

        # Wait for the user to drag the mouse
        end_x, end_y = pyautogui.position()
        print(f"Ending position: ({end_x}, {end_y})")

        # Calculate the region coordinates
        x = min(start_x, end_x)
        y = min(start_y, end_y)
        width = abs(end_x - start_x)
        height = abs(end_y - start_y)

        print(f"Selected region: x={x}, y={y}, width={width}, height={height}")
        return x, y, width, height
    except Exception as e:
        print(f"Error selecting region: {e}")
        return None

# Function to extract text from a specific screen region
def extract_text_from_region(x, y, width, height):
    """
    Extracts text from a specified screen region using OCR.
    """
    try:
        # Capture the screen region
        screenshot = ImageGrab.grab(bbox=(x, y, x + width, y + height))
        screenshot_path = os.path.join(SCREENSHOT_FOLDER, f"screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        screenshot.save(screenshot_path)  # Save screenshot for debugging
        print(f"Screenshot saved: {screenshot_path}")

        # Extract text using Tesseract OCR
        text = pytesseract.image_to_string(screenshot)
        print("Text extracted successfully.")
        return text
    except Exception as e:
        print(f"Error extracting text: {e}")
        return None

# Function to save extracted text to a file
def save_extracted_text(text):
    """
    Saves the extracted text into a file in the 'extracted_text' folder.
    """
    try:
        # Generate a timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(EXTRACTED_TEXT_FOLDER, f"extracted_{timestamp}.txt")

        # Save the text to the file
        with open(filename, "w", encoding="utf-8") as file:
            file.write(text)
        
        print(f"Extracted text saved to: {filename}")
        return filename
    except Exception as e:
        print(f"Error saving extracted text: {e}")
        return None

# Function to write text to Excel
def write_to_excel(text, excel_file, sheet_name, start_row, start_col):
    """
    Writes extracted text to an Excel file in the specified position.
    """
    try:
        # Load or create the Excel workbook
        if os.path.exists(excel_file):
            workbook = openpyxl.load_workbook(excel_file)
        else:
            workbook = openpyxl.Workbook()
        
        # Select or create the sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        
        # Split text into rows and columns
        rows = text.split('\n')
        for i, row in enumerate(rows):
            cells = row.split('\t')
            for j, cell in enumerate(cells):
                sheet.cell(row=start_row + i, column=start_col + j, value=cell.strip())
        
        # Save the workbook
        workbook.save(excel_file)
        print(f"Text written to Excel: {excel_file}")
    except Exception as e:
        print(f"Error writing to Excel: {e}")

# Function to start the process
def start_process():
    """
    Main function to extract text and write to Excel.
    """
    try:
        # Select the region
        region = select_region()
        if not region:
            messagebox.showerror("Error", "Region selection failed. Please try again.")
            return
        
        x, y, width, height = region

        # Extract text from the region
        text = extract_text_from_region(x, y, width, height)
        if not text:
            messagebox.showerror("Error", "No text extracted. Please try again.")
            return
        
        # Save extracted text to a file
        save_extracted_text(text)

        # Write text to Excel
        write_to_excel(text, EXCEL_FILE, SHEET_NAME, start_row=2, start_col=2)
        messagebox.showinfo("Success", "Text copied to Excel successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Create a popup window
def create_popup():
    """
    Creates a GUI popup to start the process.
    """
    root = tk.Tk()
    root.title("Text Extractor")
    root.geometry("300x100")
    
    start_button = tk.Button(root, text="Start Process", command=start_process)
    start_button.pack(pady=20)
    
    root.mainloop()

# Entry point
if __name__ == "__main__":
    create_popup()